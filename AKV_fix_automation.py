import os
import pandas as pd
import openpyxl as op
from openpyxl.styles import PatternFill
import warnings
import re
import time

warnings.simplefilter(action='ignore')  # Disable warnings for clean output

# Function to clean column names
def clean_column_name(name):
    if isinstance(name, str):
        name = re.sub(r"Unnamed:\s*\d+_level_\d+", '', name)
        name = re.sub(r'\n+','', name)
        return name.strip()
    return name

# Function to determine the 'КВ' (coefficient) based on conditions
def determine_kv(row):
    bt = row['Базовый тариф']
    if bt == 15756:
        per = row['Перестрахование']
        if per == 0:
            return 0.01
        elif per == 1:
            return 0
    elif bt == 7535:
        branch = row['Филиал']
        date = pd.to_datetime(row['Срок действия договора подписание'])
        # Branch and date-specific logic for setting 'КВ'
        if branch in ["Тула", "Кемерово", "Тюмень", "Смоленск"]:
            if pd.Timestamp('2023-11-13') <= date <= pd.Timestamp('2024-05-31'):
                return 10
            elif date >= pd.Timestamp('2024-06-01'):
                return 5
            elif date < pd.Timestamp('2023-11-13'):
                return 0
        # More branch-specific conditions...
        return 0
    return 0

# Loop through all files with '.xlsx' extension
for file in [f for f in os.listdir() if f.endswith('.xlsx')]:
    print(f"Processing file: {file} -->\n")
    try:
        # Read Excel file and clean columns
        df = pd.read_excel(file, header=[12, 13])
        df.columns = [" ".join(col).strip() for col in df.columns.values]
        df.columns = [clean_column_name(col) for col in df.columns]
        print("Data has already been read -->\n")
            
        # Filter dataframe based on conditions
        filter_df = df[
        ((df['Статус договора страхования'] == 'Вступил в силу') & 
         (df['Превышение КВ согласовано с андеррайтерами'] == 'нет') & 
         (df['Перестрахование'] == 1) & 
         (df['Базовый тариф'] == 7535) & 
         (df['ЦФО'].str.contains('_РОЗН', case=False))
        ) |
        ((df['Статус договора страхования'] == 'Вступил в силу') & 
         (df['Превышение КВ согласовано с андеррайтерами'] == 'нет') & 
         (df['Базовый тариф'] == 15756)
        ) |
        ((df['Статус договора страхования'] == 'Вступил в силу') & 
         (df['Превышение КВ согласовано с андеррайтерами'] == 'нет') & 
         (df['Перестрахование'] == 1))
        ].copy()

        # Apply 'КВ' determination logic
        filter_df['КВ'] = filter_df.apply(determine_kv, axis=1)
        filter_df['Договор страхования Номер'] = filter_df['Договор страхования Номер'].astype(str).str.zfill(10)
        outpyt_df = filter_df[['Филиал', 'Договор страхования Номер', 'КВ']]
        print("Data to change has already been found -->\n")

        # Prepare data for modification in Excel
        new_df = outpyt_df
        new_df['Договор страхования Номер'] = new_df['Договор страхования Номер'].astype(str).str.zfill(10)
        contracts = new_df['Договор страхования Номер'].tolist()
        kv_dict = new_df.set_index('Договор страхования Номер')['КВ'].to_dict()
        number_dog = list(kv_dict.keys())
        
        # Open Excel file and modify relevant cells
        wb = op.load_workbook(file)
        ws = wb.active
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in number_dog:
                    right_cell = ws.cell(row=cell.row, column=cell.column+24)
                    right_cell.value = kv_dict[cell.value]
                    cell.fill = red_fill
        wb.save(file)
        print('The data has already been changed!\n\n')
    except Exception as e:
        print(f"File {file} doesn't have the format required for processing\n")
        continue

print("That's all!!!")
time.sleep(3)